import os
import pytest
import openpyxl
from src import export_excel

class TestDataIngestion:
    @classmethod
    def setup_class(cls):
        """
        Setup logic shared by all tests in the class.
        Runs once before any test methods are executed
        """

        # Load Test Workbook
        print(os.getcwd())
        # C:\Users\urban\Documents\GitHub\CTS_Creation\tests\CTS_Test.xlsx
        cls.test_workbook = openpyxl.load_workbook(os.path.join("tests", "CTS_Test.xlsx"))
        cls.test_worksheet = cls.test_workbook["MAP or COFA"]


    def test_get_column_letter(self):
       
        expected_col_letter = "B"
        col_letter = export_excel.get_column_letter(self.test_worksheet, "LAST NAME")

        assert col_letter == expected_col_letter

        with pytest.raises(ValueError):
            col_letter = export_excel.get_column_letter(self.test_worksheet, "NOT IN SHEET")

    def test_add_protection(self):
       
        password = "test"
        col_to_unlock = ["LAST NAME"]

        export_excel.protection_handler(self.test_workbook, col_to_unlock, password=password, row_range=5)

        for row in range(2, 6):                 
            cell = self.test_worksheet[f"B{row}"]  # Access the cell in column B
            
            assert cell.protection.locked is False

        protected_cell = self.test_worksheet["A1"]

        assert protected_cell.protection.locked is True
