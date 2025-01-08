"""
Module: modify_excel
Description: This module handles the implementation of Data Validation and formatting
             for the generated Excel spreadsheet from export_excel.py. The module will
             make use of the openpyxl library which provides methods for modifying an Excel
             workbook. We use a class instantiation of an openpyxl.workbook and create
             custom methods to create the spreadsheet with the desired formatting and
             data validation.

Author: Urban Halpern
Date: 2024-12-26
"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


class CustomSpreadsheet:
    def __init__(self, filepath: str = None, row_range: int = 1000):
        """
        Initializes a CustomSpreadsheet object to modify or create a workbook.

        Args:
            filepath (str): The path to an existing Excel file to load.
                            If not provided, a new workbook will be created.
            row_range: The number of rows that should be formatted

        Attributes:
            workbook (Workbook): The openpyxl Workbook object representing the loaded or newly created workbook.
            sheet (Worksheet): The active worksheet in the workbook.

        Raises:
            FileNotFoundError: If the filepath is provided but the file does not exist.
        """

        if filepath:
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"The file '{filepath}' does not exist.")
            try:
                self.workbook = load_workbook(filepath)
            except Exception as e:
                raise IOError(f"An error occurred while loading the file: {e}")  # __init__ terminates immediately

        self.sheet = None
        self.range = row_range

    # Simple function to set the active sheet using the sheet name
    def set_sheet(self, sheet_name: str):
        self.sheet = self.workbook[sheet_name]
      

    # Applies the specified background color to the header row
    def set_header_style(self, color_code: str = "4472c4", font_size=8, font_name="Calibri"):

        # Define styles for fill, font, and alignment
        header_fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        fontStyle = Font(size=font_size, name=font_name, bold=True)
        allign = Alignment(horizontal="center", vertical="center")

        # Apply fonts to cells
        for cell in self.sheet[1]:
            cell.fill = header_fill
            cell.font = fontStyle
            cell.alignment = allign

    # Sets the specified height (in pixels) of the header row
    def set_header_height(self, header_row_height: float = 22.9):
       self.sheet.row_dimensions[1].height = header_row_height

    def set_column_width(self, multiplier: float = 1.2):
        """
        Sets the column widths based on the column name. The function will iterate
        through all the columns with data in them. The column width will be modified
        according to the cell with the longest value. This will fit each column length
        so that everything is displayed neatly and no information is cut off

        Args:
            multiplier (float): Controls extra whitespace to add to the column widths

        """

        # TODO: Have a max width cut off point for long cell values

        for col in self.sheet.columns:
            max_length = 8 # Set a default width
            column = col[0].column_letter
            
            # Find cell in column with longest value
            for cell in col:
                
                # if cell is datetime column value, truncate value to just consider date
                if isinstance(cell.value, datetime):
                    # Retain only the date part
                    cell.value = cell.value.date()

                # If cell is not empty, check for width
                if cell.value is not None:
                    # If the current cell is larger than the cell width, increase the max width
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

            # Adjust cell width to width of max value
            adjusted_width = (max_length + 1) * multiplier
            self.sheet.column_dimensions[column].width = adjusted_width
        

    def set_alternating_fill(self, first_color: str = "d9e1f2", second_color: str = "b4c6e7"):
        """
        Creates an alternating pattern on the spreadsheet of filled rows and non-filled rows.
        The row colors will be formatted up to the defined stop_row with the specified color.

        Args:
            first_color (str): to fill all the even rows
            second_color (str): to fill all the odd rows

        """

        # Define fills and border style
        fill_one = PatternFill(start_color=first_color, end_color=first_color, fill_type="solid")
        fill_two = PatternFill(start_color=second_color, end_color=second_color, fill_type="solid")
        thin_border = Border(left=Side(style='thin', color='595959'),
                     right=Side(style='thin', color='595959'),
                     top=Side(style='thin', color='595959'),
                     bottom=Side(style='thin', color='595959'))
      
        for row in self.sheet.iter_rows(min_row=2, max_row=self.range):
            # retrieve the row number of the current row object and check if even or odd
            # fill even rows with fill one and odd row with fill two
            if row[0].row % 2 == 0:
                fill = fill_one
            else:
                fill = fill_two

            # Apply the fill to all the cells in the row
            for cell in row:
                cell.fill = fill
                cell.border = thin_border

    def set_number_of_columns(self, num_columns: int = 22):
        """
        This method sets the number of columns to display on the spreadsheet. It will
        hide every column past the num_columns specified. It will have to iterate all
        the way up to the max column for Excel spreadsheets: 16285

        Args:
            num_columns (int): number of columns to display
        """

        for col in range(num_columns, 16385):

            # return cell object based on given coordinates and return column_letter property
            col_letter = self.sheet.cell(row=1, column=col).column_letter

            # set column to hidden
            self.sheet.column_dimensions[col_letter].hidden = True

    def add_style_format(self, format_str: str, col: str):
        """
        Sets a spreadsheet column to have a format equivalent to a custom format in Excel
        Will allow certain columns to automatically convert recognized inputs to format.
        Unrecognized inputs will be caught by data validation.

        Args:
            format_str (str): format string to apply to column
            col (str): column letter to apply formatting to
        """

        for cell in self.sheet[col]:
            cell.number_format = format_str

    def add_value_formula(self, value_formula: str, col: str,):
        """
        Adds a value formula to a specified column in the sheet object. A value formula
        is a formula applied to a cell that calculates the value in that cell.
        Example formula: =FLOOR($M{row}*0.17,0.01)

        Args:
            value_formula (str): formula to apply to cell
            col (str): column letter representing the column to add formula to
        """
        # Check for the '{row}' placeholder using regex
        if not re.search(r"\{row\}", value_formula):
            raise ValueError(f"The provided formula: {value_formula} does not contain a 'row' placeholder.")

        # Start iterating at 2 to skip the header row. Excel rows start at 1
        for row_num in range(2, self.range + 1):

            # Apply the formula to each row in the column.
            # The keyword argument row represents placeholders in the formula string that will be replaced by row_num
            self.sheet[f'{col}{row_num}'] = value_formula.format(row=row_num)
            
    def add_data_validation(self, formula: str,
                            col_to_validate: str,
                            error_message: str = "Error, the data you entered violates the data validation rules rule set for this cell."):
        """
        Adds a defined datavalidation object to the sheet based on a defined formula 
        The function will apply the data validation to all rows in the columns using self.range

        Args:
            formula (str): data validation formula to add to sheet
            col_to_validate (str): the letter of the column to apply the validation to
            error_message (str): error message to display when data validation is violated

        """

        # Define a data validation object
        dv = DataValidation(type="custom",
                            formula1=formula,
                            showErrorMessage=True)

        dv.error = error_message
        # Apply the validation to range (EX: M2:M1000)
        add_str = f'{col_to_validate}2:{col_to_validate}{self.range}'

        dv.add(add_str) 

        # Add validation object to the sheet
        self.sheet.add_data_validation(dv)

    def add_conditional_formatting(self, formula: str, col_to_validate: int):
        """
        Adds a defined rule object to the sheet. The rule object should
        already be defined and the column to apply the formatting rule
         should be specified. The function will apply the formatting rule
         to all cells in the column

        Args:
            formula (str): conditional formatting formula used to creat Rule object
            col_to_validate (int): the integer value of the column to apply the validation to


        """

        # Define conditional formatting rule to add
        yellow_highlight = PatternFill(start_color='ffff00',
               end_color='ffff00',
               fill_type='solid')
        rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=yellow_highlight)

        add_str = f'{col_to_validate}2:{col_to_validate}{self.range}'

        # Add conditional formatting to sheet
        self.sheet.conditional_formatting.add(add_str, rule)


    def get_column_letter(self, column_name: str) -> str:
        """
        Helper function to find the column letter from a specified column_name. If no matching
        header is found, will return None.

        Args:
            column_name (str): column name to extract header from

        Returns:
            column_letter (str): column letter associated with header.
        """

        column_letter = None

        # Iterate through all of the columns in the spreadsheet returning cells only up to row 2 for each column
        for col in self.sheet.iter_cols(max_row=2):

            # Extract the first row cell
            header_cell = col[0]
         
            # If header cell name found, return column letter
            if header_cell.value == column_name:
                return header_cell.column_letter
        
        # Raise error if the column was not found in the sheet.
        if column_letter is None:
            raise ValueError(f'Specified Column: {column_name} not found in sheet.')


def formatting_handler(workbook: CustomSpreadsheet, validation_format_dict: dict) -> None:
    """
        This will ingest all of the data validation and conditional formatting
        definitions for each of the columns into a CustomSpreadsheet object.

        Args:
            workbook (CustomSpreadsheet): The workbook to add the data validation rules to
            validation__format dict (dict): Holds the defined data validation and conditional formatting rules for each column.
                                            {header_name: {validation_params}}

        Returns:
            None
    """

    # For each header in the validation_format_dict, add associated formatting
    for header in validation_format_dict:

        # Get the column letter from the header_name
        col = workbook.get_column_letter(header)
      
        # For each of the following:
        # Check if key exists and add associated format to column

        # Data Validation
        validation = validation_format_dict[header].get("data_validation")
        error_message = validation_format_dict[header].get("error_msg")
        if validation is not None:
            workbook.add_data_validation(validation, col, error_message=error_message)

        # Conditional Formatting
        conditional_format_formula = validation_format_dict[header].get("conditional_format_formula")
        if conditional_format_formula is not None:
            workbook.add_conditional_formatting(conditional_format_formula, col)

        # Style Formatting
        style_format = validation_format_dict[header].get("style_format")
        if style_format is not None:
            workbook.add_style_format(style_format, col)

        # Value Formatting
        value_formula = validation_format_dict[header].get("value_formula")
        if value_formula is not None:
            workbook.add_value_formula(value_formula, col)