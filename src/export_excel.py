"""
Module: export_excel
Description: This module handles the creation of a claims transmittal spreadsheet
             from a prepared pandas dataframe. It also provides funcitonality for 
             ingesting data into a spreadsheet template.

Author: Urban Halpern
Original Creation: 2024-12-24
Latest Revision: 2025-01-17
"""

import os
import pandas as pd
import openpyxl
import openpyxl.workbook
from openpyxl import load_workbook
from openpyxl.styles import Protection, Alignment


def get_format(cell: openpyxl.cell.cell.Cell, validation_format_dict: dict, header: str) -> None:
    """
    Sets the number formatof a cell from the validation_format_dict

    Args:
        sheet: (openpyxl.worksheet.worksheet): spreadsheet object for accessing header cell
        cell (openpyxl.cell.cell.Cell): cell to get formatting for
        validation_format_dict (dict): dictionary that holds formatting for each column
        header (str): header name to lookup in validation_format_dict

    """

    # Look for the header in the validation_format_dict
    format_rules = validation_format_dict.get(header)

    # Apply style formatting if the header and style_format field is found
    if format_rules is not None:
        style_format = format_rules.get("style_format")
        alignment = format_rules.get("alignment")

        # Style Formatting
        if style_format is not None:
            cell.number_format = style_format

        # Alignment
        if alignment is not None:
            cell.alignment = Alignment(horizontal=alignment)


def insert_into_template(final_df: pd.DataFrame, validation_format_dict: dict) -> openpyxl.workbook.workbook.Workbook:
    """
    Inserts data into the template spreadsheet using data from the final_df by column

    Args:
        final_df (pandas.dataframe): dataframe which holds transformed data from SQL query
        validation_format_dict (dict): dictionary that holds formatting for each column
        workbook_name (str): name of workbook that will be saved after data is ingested

    Returns:
        workbook (openpyxl.workbook.workbook.Workbook): workbook with ingested data

    """

    # Get parent dir of repo to access generated_sheets dir
    parent_dir = os.path.abspath(os.path.join(os.getcwd()))

    # Access the template file
    template_file_path = os.path.join(parent_dir, 'CTS_Example_Template.xlsx')

    workbook = load_workbook(template_file_path)
    sheet = workbook["MAP or COFA"]

    # Iterate though the columns in the dataframe
    for col_name in final_df.columns:

        # Find the column in the sheet
        col_letter = get_column_letter(sheet, col_name)
        col_data = final_df[col_name]

        # iterate through the rows in the column, skipping the header cell and using one-based indexing
        for row_idx, value in enumerate(col_data, start=2):

            cell = sheet[f"{col_letter}{row_idx}"]

            # Skip cells with formulas
            if not cell.data_type == "f":
                cell.value = value

                get_format(cell, validation_format_dict, col_name)

    return workbook


def save_workbook(workbook: openpyxl.workbook.Workbook, workbook_name: str = "CTS_Insert_Example.xlsx") -> None:
    """
    Saves the workbook to the specified path and checks if file already exists

    Args:
        workbook (openpyxl.workbook.Workbook): The workbook object to save
        worbook_name (str): Name of file

    """
    # Get parent dir of repo to access generated_sheets dir
    parent_dir = os.path.abspath(os.path.join(os.getcwd()))
    sheets_directory = os.path.join(parent_dir, 'generated_sheets')

    # Define path to save workbook and raise error if wb with same name exists
    save_path = os.path.join(sheets_directory, workbook_name)
    if os.path.exists(save_path):
        raise FileExistsError(f'The file already exists: {save_path}')

    # Save workbook
    workbook.save(save_path)
    print(f'Sheet saved to {workbook_name} at {save_path}') 

    return save_path


def protection_handler(workbook: openpyxl.workbook.Workbook, cols_to_unprotect: list,
                       password: str = "test", row_range: int = 50) -> None:
    """
    Un-protects columns that do not need protection. Should only unprotect
    number of rows equal to the SQL query.

    Args:
        workbook (openpyxl.workbook.Workbook): The workbook with columns to unprotect
        cols_to_unprotect (list): List of column headers to unprotect
        password (str): password to unlock the sheet
        range (int): range of cells in column to unprotect

    """

    # Protect all cells and set password
    sheet = workbook["MAP or COFA"]
    sheet.protection.enable()
    sheet.protection.password = password

    for column in cols_to_unprotect:
        col_letter = get_column_letter(sheet, column)

        unlock_column(sheet, col_letter, row_range)


def get_column_letter(sheet: openpyxl.worksheet.worksheet.Worksheet, column_name: str) -> str:
    """
    Helper function to find the column letter from a specified column_name. If no matching
    header is found, will return None.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): worksheet object to extract column letter from
        column_name (str): column name to extract header from

    Returns:
        column_letter (str): column letter associated with header.
    """

    column_letter = None

    # Iterate through all of the columns in the spreadsheet returning cells only up to row 2 for each column
    for col in sheet.iter_cols(max_row=2):

        # Extract the first row cell
        header_cell = col[0]
      
        # If header cell name found, return column letter
        if header_cell.value == column_name:
            return header_cell.column_letter
  
    # Raise error if the column was not found in the sheet.
    if column_letter is None:
        raise ValueError(f'Specified Column: {column_name} not found in sheet.')


def unlock_column(sheet: openpyxl.worksheet.worksheet.Worksheet, column_to_unlock: str, row_range: int):
    """
    This method unlocks the cells in a column to allow user entry
    assuming the sheet was already set to protection mode.
    The header cell will remain locked for each column.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): sheet to unlock columns in
        column_to_unlock (str): letter of the column to unlock
        range (int): Range of cells to unlock
    """

    # iterate through cells in the column, skipping the header cell
    for cell in sheet[column_to_unlock][1:row_range]:
        cell.protection = Protection(locked=False)  # Unlock the cell
